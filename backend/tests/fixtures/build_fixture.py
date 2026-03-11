"""Generates sample_import.xlsx — the golden file fixture for A3.

Run this script once to regenerate the fixture:
    docker compose exec -e PYTHONPATH=/app backend python tests/fixtures/build_fixture.py

The generated file is committed to the repo. This script is the source of truth
for its contents. If you change the data, regenerate AND update expected_output.json.
"""

from pathlib import Path
import openpyxl

OUTPUT_PATH = Path(__file__).parent / "sample_import.xlsx"

# ──────────────────────────────────────────────
# Data definitions
# ──────────────────────────────────────────────

METADATA = [
    ("report_month", "2025-09-01"),
    ("version_type", "current"),
    ("source_type", "excel"),
    ("actual_flags", "true,true,true"),
    ("n12m_months", "2025-10-01,2025-11-01,2025-12-01"),
]

DELIVERY_COUNTS = [
    ("p1", 2),
    ("p2", 3),
    ("p3", 4),
    ("p4", 0),
    ("p5", 0),
]

# Raw input N12M rows only (is_calculated=false).
# Costs are stored as POSITIVE values (v2 convention).
# Columns: section, line_item, display_name, is_calculated, sort_order, Oct, Nov, Dec
N12M_ROWS = [
    ("revenue",      "gross_revenue",           "Gross Revenue",             False,  1,  60000,  40000,  20000),
    ("revenue",      "sales_costs",             "Selling Costs",             False,  2,   3000,   2000,   1000),
    ("direct_costs", "primary_build",           "Primary Build",             False,  3,  30000,  20000,  10000),
    ("direct_costs", "contingency",             "Contingency",               False,  4,      0,      0,      0),
    ("overheads",    "marketing",               "Marketing",                 False,  5,   1500,   1000,    500),
    ("overheads",    "admin_overheads",         "Admin Overheads",           False,  6,      0,      0,      0),
    ("capex",        "infrastructure",          "Infrastructure",            False,  7,  15000,  10000,   5000),
    ("capex",        "civil_works",             "Civil Works",               False,  8,      0,      0,      0),
    ("capex",        "landscaping",             "Landscaping",               False,  9,      0,      0,      0),
    ("capex",        "amenities",               "Amenities",                 False, 10,      0,      0,      0),
    ("capex",        "professional_fees",       "Professional Fees",         False, 11,   3000,   2000,   1000),
    ("capex",        "regulatory_fees",         "Regulatory Fees",           False, 12,      0,      0,      0),
    ("capex",        "other",                   "Other",                     False, 13,      0,      0,      0),
    ("capex",        "contingency_civil_works", "Contingency Civil Works",   False, 14,      0,      0,      0),
    ("capex",        "contingency_amenities",   "Contingency Amenities",     False, 15,      0,      0,      0),
    ("capex",        "ancillary_build_capex",   "Ancillary Build Capex",     False, 16,      0,      0,      0),
]

# Phase comparison — whole-of-project current values.
# Only line items with non-zero values are included — 6 items × 6 phases = 36 rows.
# Columns: line_item, p1, p2, p3, p4, p5, total
PHASE_COMPARISON_ROWS = [
    ("gross_revenue",    100000, 150000, 250000, 0, 0, 500000),
    ("sales_costs",        5000,   7500,  12500, 0, 0,  25000),
    ("primary_build",     50000,  75000, 125000, 0, 0, 250000),
    ("marketing",          5000,   7500,  12500, 0, 0,  25000),
    ("infrastructure",    20000,  30000,  50000, 0, 0, 100000),
    ("professional_fees", 10000,  15000,  25000, 0, 0,  50000),
]


# ──────────────────────────────────────────────
# Build workbook
# ──────────────────────────────────────────────

def build():
    wb = openpyxl.Workbook()

    # Sheet 1: metadata
    ws = wb.active
    ws.title = "metadata"
    ws.append(["key", "value"])
    for row in METADATA:
        ws.append(list(row))

    # Sheet 2: delivery_counts
    ws = wb.create_sheet("delivery_counts")
    ws.append(["phase", "count"])
    for row in DELIVERY_COUNTS:
        ws.append(list(row))

    # Sheet 3: n12m
    ws = wb.create_sheet("n12m")
    ws.append([
        "section", "line_item", "display_name", "is_calculated", "sort_order",
        "2025-10-01", "2025-11-01", "2025-12-01",
    ])
    for row in N12M_ROWS:
        ws.append(list(row))

    # Sheet 4: phase_comparison (current values only; budget from linked import)
    ws = wb.create_sheet("phase_comparison")
    ws.append(["line_item", "p1", "p2", "p3", "p4", "p5", "total"])
    for row in PHASE_COMPARISON_ROWS:
        ws.append(list(row))

    wb.save(OUTPUT_PATH)
    print(f"Written: {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
